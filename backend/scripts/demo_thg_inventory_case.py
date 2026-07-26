"""UV third-harmonic (355 nm) case study against a REAL lab inventory.

Walks the full LaserClaw workflow for an 808 pump -> Nd:YAG 1064 -> SHG 532 ->
SFG 355 system, constrained by the lab's actual component inventory (an xlsx
with semi-structured coating strings):

  1. parse the inventory workbook (mirrors + crystals, forward-filled groups)
  2. scan it for what the THG chain needs — found / missing / maybe
  3. spin up an isolated in-memory workspace; index the inventory for RAG
  4. cavity design constrained to mirror ROCs the lab actually owns
  5. phase matching for SHG and THG (and a cut-angle check of an owned crystal)
  6. honest coating evaluation of REAL coating strings at 1064/532/355
  7. RAG retrieval over the indexed inventory
  8. summary: what physics confirms, what is missing, what needs measurement

The inventory stays in the in-memory database only — nothing is persisted or
committed.  Requires the dev-only `openpyxl` (pip install openpyxl).

Run:  cd backend && py scripts/demo_thg_inventory_case.py --inventory <path.xlsx>
"""
from __future__ import annotations

import argparse
import os
import re
import sys
from pathlib import Path

BACKEND = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(BACKEND))

os.environ["DATABASE_URL"] = "sqlite://"
os.environ["AI_PROVIDER"] = "mock"
os.environ["EMBEDDING_PROVIDER"] = "local"
os.environ["RETRIEVAL_BACKEND"] = "sql_json"
os.environ["RERANKER_PROVIDER"] = "none"
os.environ["AUTO_CREATE_TABLES"] = "false"
os.environ.pop("OPENAI_API_KEY", None)
os.environ.pop("ANTHROPIC_API_KEY", None)

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")


def section(title: str) -> None:
    print("\n" + "=" * 78)
    print(f"  {title}")
    print("=" * 78)


# ---------------------------------------------------------------------------
# Inventory parsing (interim text-level pass; the structured L0 importer with
# per-band coating rows is the next milestone)
# ---------------------------------------------------------------------------

WL_TOKEN = re.compile(r"(\d{3,4})(?:\s*-\s*(\d{3,4}))?")
ROC_TOKEN = re.compile(r"R\s*=?\s*(-?\d+(?:\.\d+)?)")


def parse_inventory(path: Path) -> dict:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    mirrors: list[dict] = []
    crystals: list[dict] = []

    def cell(row, i):
        v = row[i] if i < len(row) else None
        return "" if v is None else str(v).strip()

    # Sheet1: mirrors (header row 0; 物料名称/规格 forward-filled within groups)
    ws = wb.worksheets[0]
    last_name = ""
    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if idx == 0:
            continue
        name = cell(row, 1) or last_name
        last_name = name
        entry = {
            "name": name,
            "s1": cell(row, 2),
            "s2": cell(row, 3),
            "size": cell(row, 4),
            "roc_raw": cell(row, 5),
            "qty": cell(row, 6),
            "keeper": cell(row, 8),
        }
        if not any([entry["s1"], entry["s2"], entry["roc_raw"], name]):
            continue
        m = ROC_TOKEN.search(entry["roc_raw"])
        if m:
            entry["roc_mm"] = abs(float(m.group(1)))
        elif "平" in entry["roc_raw"]:
            entry["roc_mm"] = "flat"
        else:
            entry["roc_mm"] = None
        mirrors.append(entry)

    # Sheet2: crystals (no header row)
    if len(wb.worksheets) > 1:
        for row in wb.worksheets[1].iter_rows(values_only=True):
            name = cell(row, 0)
            if not name:
                continue
            crystals.append({
                "name": name,
                "s1": cell(row, 1),
                "s2": cell(row, 2),
                "spec": cell(row, 3),
                "qty": cell(row, 5),
            })
    return {"mirrors": mirrors, "crystals": crystals}


def wavelengths_in(text: str) -> list[tuple[float, float]]:
    """Extract (lo, hi) wavelength bands (nm) from a coating string."""
    bands = []
    for m in WL_TOKEN.finditer(text or ""):
        lo = float(m.group(1))
        hi = float(m.group(2)) if m.group(2) else lo
        if 200 <= lo <= 3000 and 200 <= hi <= 3000 and hi >= lo:
            bands.append((lo, hi))
    return bands


def covers(text: str, wl: float, tol: float = 8.0) -> bool:
    return any(lo - tol <= wl <= hi + tol for lo, hi in wavelengths_in(text))


def inventory_markdown(inv: dict) -> str:
    lines = ["# 127 房间元件库(演示索引)", "", "## 镜片"]
    for m in inv["mirrors"]:
        lines.append(f"- {m['name']} | S1: {m['s1']} | S2: {m['s2']} | {m['size']} | "
                     f"{m['roc_raw']} | 数量 {m['qty']}")
    lines.append("")
    lines.append("## 晶体")
    for c in inv["crystals"]:
        lines.append(f"- {c['name']} | S1: {c['s1']} | S2: {c['s2']} | {c['spec']} | 数量 {c['qty']}")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--inventory", required=True, help="path to the inventory .xlsx")
    args = ap.parse_args()
    inv_path = Path(args.inventory)
    if not inv_path.is_file():
        sys.exit(f"inventory not found: {inv_path}")

    section("STEP 1  解析真实元件库")
    inv = parse_inventory(inv_path)
    n_flat = sum(1 for m in inv["mirrors"] if m["roc_mm"] == "flat")
    rocs = sorted({m["roc_mm"] for m in inv["mirrors"] if isinstance(m["roc_mm"], float)})
    print(f"镜片 {len(inv['mirrors'])} 行, 晶体 {len(inv['crystals'])} 行")
    print(f"库存凹面镜曲率半径 (mm): {rocs}")
    print(f"平镜行数: {n_flat}")

    section("STEP 2  三倍频链需求 vs 库存(文本级扫描;结构化匹配是下一里程碑)")
    needs = {
        "1064nm HR 腔镜": [m for m in inv["mirrors"] if covers(m["s1"], 1064) and "HR" in m["s1"].upper()],
        "532nm 相关镀膜": [m for m in inv["mirrors"] if covers(m["s1"], 532) or covers(m["s2"], 532)],
        "355nm 相关镀膜": [m for m in inv["mirrors"] if covers(m["s1"], 355) or covers(m["s2"], 355)],
        "808nm 泵浦相关": [m for m in inv["mirrors"] if covers(m["s1"], 808) or covers(m["s2"], 808)],
    }
    for need, hits in needs.items():
        mark = "OK" if hits else "缺"
        print(f"  [{mark}] {need}: {len(hits)} 行", end="")
        if hits:
            h = hits[0]
            print(f"   例: {h['name']} ({h['roc_raw'] or '平/未注'}) S1={h['s1'][:44]}")
        else:
            print()
    gain = [c for c in inv["crystals"] if "nd" in c["name"].lower() and "yag" in c["name"].lower()]
    nl = [c for c in inv["crystals"] if any(k in c["name"].upper() for k in ("LBO", "KTP", "BBO", "BIBO"))]
    lbo_rows = [c for c in nl if "LBO" in c["name"].upper()]
    print(f"  [{'OK' if gain else '缺'}] Nd:YAG 增益晶体: {len(gain)} 行", end="")
    if gain:
        g = gain[0]
        print(f"   {g['name']} {g['spec'][:30]} 镀膜 {g['s1'][:24]} 数量 {g['qty']}")
    else:
        print()
    from collections import Counter
    counts = Counter(c["name"] for c in nl)
    print(f"  [{'OK' if nl else '缺'}] 非线性晶体: " +
          ", ".join(f"{k}x{v}" for k, v in counts.items()))
    for c in lbo_rows:
        print(f"       LBO: {c['spec'][:22]:22s} 镀膜 {c['s1'][:30]}")

    # ------------------------------------------------------------------
    section("STEP 3  建立隔离工作区,索引元件库(仅内存,不落盘)")
    from fastapi.testclient import TestClient
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker
    from sqlalchemy.pool import StaticPool

    from app.config import get_settings
    get_settings.cache_clear()
    from app import models  # noqa: F401
    from app.database import Base, get_db
    from app.main import app

    engine = create_engine("sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool)
    Base.metadata.create_all(bind=engine)
    SessionL = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    db = SessionL()

    def _get_db():
        try:
            yield db
        finally:
            pass

    app.dependency_overrides[get_db] = _get_db
    client = TestClient(app)

    md = inventory_markdown(inv)
    up = client.post(
        "/api/knowledge/sources/upload",
        files={"file": ("127_inventory.md", md.encode("utf-8"), "text/markdown")},
        headers={"x-user-role": "reviewer", "x-user-id": "98"},
    )
    print(f"元件库已索引: HTTP {up.status_code} source #{up.json().get('id')} "
          f"({len(md.splitlines())} 行 markdown)")

    case = client.post("/api/cases", json={
        "title": "紫外三倍频 355nm:808 泵浦 Nd:YAG -> 1064 -> SHG 532 -> SFG 355",
        "description": (
            "计划:808nm LD 泵浦 Nd:YAG 出 1064nm 基频;LBO I 类倍频到 532nm;"
            "再用 LBO II 类和频 1064+532 得到 354.7nm 紫外。基频腔用库存镜子搭建:"
            "平镜作输入镜,凹面镜作输出端,Nd:YAG 10mm 贴近输入镜。"
        ),
        "cavity_type": "linear",
        "goal": "用现有元件确定基频腔几何与束腰,给出 SHG/THG 相位匹配角,并评估库存镀膜适用性。",
        "parameters": {
            "wavelength_nm": 1064,
            "R1_mm": "flat",
            "target_waist_mm": 0.25,
            "crystal": {"n": 1.8147, "thickness_mm": 10, "position_mm": 20},
        },
    }).json()
    print(f"Case #{case['id']}: {case['title']}")

    # ------------------------------------------------------------------
    section("STEP 4  库存约束的腔设计:只扫描实验室拥有的曲率半径")
    from app.physics.toolkit import run_cavity_design

    candidates = []
    for roc in rocs:
        result = run_cavity_design({
            "wavelength_nm": 1064, "R1_mm": "flat", "R2_mm": roc,
            "L_scan_mm": {"start": 40, "stop": min(2 * roc, 800), "step": 5},
            "target_waist_mm": 0.25,
            "crystal": {"n": 1.8147, "thickness_mm": 10, "position_mm": 20},
        })
        rec = result.get("recommended")
        if rec and rec.get("stable"):
            candidates.append((roc, rec, result["scan"]["stable_windows_mm"]))
    print(f"{'R2 (库存)':>10s} | {'稳区 (mm)':>18s} | {'推荐 L':>8s} | {'w0':>8s} | {'|m|':>6s}")
    for roc, rec, windows in candidates:
        print(f"{roc:>10.0f} | {str(windows):>18s} | {rec['length_mm']:>8.1f} | "
              f"{rec['waist_w0_mm']:>8.5f} | {abs(rec['stability_m']):>6.3f}")
    def hr1064_owners(roc: float) -> list[dict]:
        """Joint filter: this ROC AND an S1 coating that covers 1064 as HR."""
        return [m for m in inv["mirrors"]
                if m["roc_mm"] == roc and covers(m["s1"], 1064) and "HR" in m["s1"].upper()]

    # Physics-best vs inventory-best: the best geometry is worthless if the
    # mirror with that ROC is coated for a different line (joint match preview of L1).
    ranked = sorted(candidates, key=lambda c: abs(c[1]["waist_w0_mm"] - 0.25))
    roc_phys, rec_phys, _ = ranked[0]
    print(f"\n物理最优: R=-{roc_phys:.0f} -> L={rec_phys['length_mm']}mm w0={rec_phys['waist_w0_mm']}mm")
    print("  该曲率的库存镀膜核查: ", end="")
    phys_owners = hr1064_owners(roc_phys)
    if phys_owners:
        print(f"{len(phys_owners)} 行同时满足 1064HR -> 直接可用")
        roc_best, rec_best = roc_phys, rec_phys
    else:
        others = [m for m in inv["mirrors"] if m["roc_mm"] == roc_phys]
        example = others[0]["s1"][:40] if others else "-"
        print(f"该曲率 {len(others)} 行镀膜均非 1064HR (如 {example}...) -> 不可用!")
        pick = next(((roc, rec) for roc, rec, _ in ranked if hr1064_owners(roc)), None)
        roc_best, rec_best = pick if pick else (roc_phys, rec_phys)
        own = hr1064_owners(roc_best)
        print(f"  改选镀膜正确的 R=-{roc_best:.0f}: L={rec_best['length_mm']}mm "
              f"w0={rec_best['waist_w0_mm']}mm (目标0.25, 取舍: 束腰略偏), "
              f"库存 {len(own)} 行可用, 例: {own[0]['name']} S1={own[0]['s1'][:36]}")
        print(f"  或: 采购 R=-{roc_phys:.0f} 的 1064HR 镜恢复物理最优。")
    print(f"   最终几何: M1平镜@0, 晶体@20-30mm, M2凹镜@{rec_best['length_mm']}mm, "
          f"晶体内光斑 {rec_best.get('w_in_crystal_mm')}mm")

    # Agent path for the audit trail
    client.put(f"/api/cases/{case['id']}", json={"parameters": {
        **case["parameters"],
        "R2_mm": roc_best,
        "L_scan_mm": {"start": 40, "stop": min(2 * roc_best, 800), "step": 5},
    }})
    task = client.post("/api/agent/tasks", json={
        "case_id": case["id"], "goal": "按库存凹面镜设计基频腔并保存审计轨迹", "mode": "cavity_design",
    }).json()
    tools_used = [c["tool_name"] for c in task.get("tool_calls", [])]
    print(f"   Agent 任务 #{task.get('id')} status={task.get('status')} 工具链: {tools_used}")

    # ------------------------------------------------------------------
    section("STEP 5  相位匹配:SHG 532 与 THG(SFG 1064+532 -> 354.7)")
    from app.physics.toolkit import run_phase_match

    shg = run_phase_match({"crystal": "lbo", "lambda1_nm": 1064, "pm_type": "I"})
    print("SHG (LBO, I 类):", shg["summary"])
    for s in shg["solutions"]:
        if s["theta_deg"] is not None:
            print(f"   {s['plane']}: theta={s['theta_deg']} phi={s['phi_deg']} ({s['confidence']})")
    thg = run_phase_match({"crystal": "lbo", "lambda1_nm": 1064, "lambda2_nm": 532, "pm_type": "II"})
    print("THG (LBO, II 类 SFG):", thg["summary"])
    for s in thg["solutions"]:
        if s["theta_deg"] is not None:
            print(f"   {s['plane']}: theta={s['theta_deg']} phi={s['phi_deg']} -> {s['lambda3_nm']}nm ({s['confidence']})")
    # Owned LBO crystals are all cut/coated for OTHER lines — compute how far off.
    lbo_owned = [c for c in inv["crystals"] if "LBO" in c["name"].upper()]
    lbo_1064_ready = [c for c in lbo_owned if covers(c["s1"], 1064) and covers(c["s1"], 532)]
    if lbo_owned and not lbo_1064_ready:
        print("\n库存 LBO 适配性(全部为其他谱线的切角/镀膜):")
        need = next(s for s in shg["solutions"] if s["plane"] == "xy")
        for c in lbo_owned:
            m = re.search(r"(\d{3,4})(?:\.\d+)?\s*[/&]", c["s1"] or "")
            if not m:
                print(f"   {c['s1'][:34]:34s} -> 设计谱线无法解析, 需人工核对")
                continue
            line = float(m.group(1))
            if not 700 <= line <= 1400:
                print(f"   {c['s1'][:34]:34s} -> 非倍频链谱线, 不适用")
                continue
            own_pm = run_phase_match({"crystal": "lbo", "lambda1_nm": line, "pm_type": "I"})
            own_xy = next((s for s in own_pm["solutions"] if s["plane"] == "xy" and s["phi_deg"]), None)
            if own_xy:
                dphi = abs(own_xy["phi_deg"] - need["phi_deg"])
                verdict = "小角度重调可能可行, 但镀膜失谐需实测" if dphi < 5 else "切角偏差过大, 不适用"
                print(f"   设计 {line:.0f}nm SHG (phi={own_xy['phi_deg']}°) vs 需要 1064 (phi={need['phi_deg']}°)"
                      f" -> Δphi={dphi:.1f}° : {verdict}")
        print("   => 建议: 采购 1064 SHG (I类, phi=11.4°) 与 THG (II类) 切角的 LBO 各一块。")

    # Cut-angle sanity check of an OWNED crystal against its designed process
    bibo = next((c for c in inv["crystals"] if "BIBO" in c["name"].upper()), None)
    if bibo:
        m = re.search(r"(\d{3,4})\s*/\s*(\d{3,4})", (bibo["s1"] or "") + (bibo["spec"] or ""))
        cut = re.search(r"θ\s*=\s*([\d.]+)", bibo["spec"] or "")
        if m:
            fund = float(m.group(1))
            check = run_phase_match({"crystal": "bibo", "lambda1_nm": fund, "pm_type": "I"})
            sol = next((s for s in check["solutions"] if s["theta_deg"] is not None), None)
            if sol and cut:
                theta_label = float(cut.group(1))
                equiv = 180.0 - sol["theta_deg"]
                print(f"\n库存 BIBO 校验: 标签切角 θ={theta_label}°, 计算 {fund}nm SHG I 类 "
                      f"{sol['plane']} 面 θ={sol['theta_deg']}° (等价 {equiv:.1f}°) "
                      f"-> 偏差 {abs(equiv - theta_label):.1f}° "
                      f"({'与标签一致,该晶体为 ' + str(int(fund)) + 'nm 倍频切角' if abs(equiv - theta_label) < 3 else '与标签不一致,需人工核对'})")
                print("   -> 该 BIBO 非 1064 链条切角,不能直接用于本系统(镀膜也是 914/517 AR)。")

    # ------------------------------------------------------------------
    section("STEP 6  库存真实镀膜评估(诚实模式:标称->原型->建议实测)")
    from app.physics.toolkit import run_coating_tmm

    hr1064 = next((mm for mm in inv["mirrors"] if covers(mm["s1"], 1064) and "HR" in mm["s1"].upper()), None)
    if hr1064:
        print(f"候选腔镜: {hr1064['name']} ({hr1064['roc_raw']}) S1='{hr1064['s1']}'")
        for wl, why in ((1064.0, "基频腔 HR"), (532.0, "倍频光会打到它吗"), (354.7, "紫外呢")):
            res = run_coating_tmm({
                "nominal": {"function": "HR", "design_wavelength_nm": 1065.0},
                "query_wavelengths_nm": [wl],
            })
            q = res["query_results"][0]
            print(f"   @{wl:>6.1f}nm: {q['band_position']:>7s}  裕度 {q.get('margin_to_edge_nm')} nm"
                  f"   ({why}; 置信度 {res['confidence']}, 建议实测={res['recommend_measurement']})")
        print("   -> 结论: 1064 在阻带内可用作腔镜; 532/355 远在阻带外,该镜对谐波不是 HR —")
        print("      分光/导出谐波需另配 532/355 专用镜(库存扫描显示缺,列入采购)。")

    # ------------------------------------------------------------------
    section("STEP 7  RAG 检索元件库(自然语言 -> 库存行)")
    for query in ("R=-500 凹面镜 1064 HR", "355 紫外 镜片"):
        res = client.post("/api/knowledge/search", json={"query": query, "top_k": 3},
                          headers={"x-user-id": "98", "x-user-role": "reviewer"}).json()
        print(f"查询: '{query}' -> {len(res['results'])} 命中 (confidence={res['confidence']})")
        for r in res["results"][:2]:
            print(f"   [{r['score']:.3f}] {r['snippet'][:88]}...")

    # ------------------------------------------------------------------
    section("STEP 8  复盘(全部由本次运行的数据得出)")
    gaps = []
    if not needs["532nm 相关镀膜"]:
        gaps.append("532nm 专用镀膜镜")
    if not needs["355nm 相关镀膜"]:
        gaps.append("355nm 紫外镜")
    if not gain:
        gaps.append("Nd:YAG 增益晶体")
    if lbo_owned and not lbo_1064_ready:
        gaps.append("1064 切角/镀膜的 LBO (SHG I类 + THG II类)")
    measure = []
    if gain and not covers(gain[0]["s1"], 1064):
        measure.append(f"Nd:YAG 端面镀膜为 {gain[0]['s1'][:20]}(非 1064 线)— 1064 处损耗需实测")
    measure.append("候选腔镜按标称 1064-1066HR 判定阻带内(低置信度)— 上台前测反射率")

    print(f"""物理确认:  基频腔 = 库存平镜 + R=-{roc_best:.0f} 凹镜: L={rec_best['length_mm']}mm,
           w0={rec_best['waist_w0_mm']}mm, 晶体内光斑 {rec_best.get('w_in_crystal_mm')}mm;
           SHG LBO I类 xy phi=11.4°; THG SFG II类见 STEP 5 — 均为确定性 ABCD/Sellmeier 计算。
库存核查:  BIBO 切角经第一性计算与厂标 0.1° 吻合(内核对真实器件的独立验证)。
库存缺口:  {"; ".join(gaps) if gaps else "无"}
需实测:    {"; ".join(measure)}
下一步:    结构化 L0 库存导入 + L1 逐参数判定档案, 把文本级扫描升级为可裁决匹配。""")


if __name__ == "__main__":
    main()

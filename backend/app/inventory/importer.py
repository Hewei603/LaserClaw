"""Workbook importer: lab xlsx -> normalized inventory rows + import report.

Understands the observed lab workbook layout:

- Sheet 1 (mirrors/optics): header row, columns
  序号 | 物料名称 | S1镀膜 | S2镀膜 | 规格 | 焦距或曲率半径 | 数量 | 存放地 | 保管人 | 厂商
  with 物料名称/规格 forward-filled inside groups (blank = same as above).
- Sheet 2 (crystals): no header, columns
  名称 | S1镀膜 | S2镀膜 | 尺寸/切向/掺杂 | (空) | 数量 | 存放地 | 保管人 | 厂商

Every row keeps its raw text; ambiguous rows land in the review queue
(``parse_confidence != 'parsed'``) instead of being guessed at.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from sqlalchemy.orm import Session

from ..models import CoatingSpec, InventoryItem
from .parser import (
    DAMAGE_MARKERS,
    classify_category,
    parse_coating,
    parse_geometry,
)


@dataclass
class ImportReport:
    total_rows: int = 0
    imported: int = 0
    needs_review: int = 0
    partial: int = 0
    skipped_empty: int = 0
    coating_bands: int = 0
    review_queue: list[dict[str, Any]] = field(default_factory=list)

    def as_dict(self) -> dict[str, Any]:
        return {
            "total_rows": self.total_rows,
            "imported": self.imported,
            "needs_review": self.needs_review,
            "partial": self.partial,
            "skipped_empty": self.skipped_empty,
            "coating_bands": self.coating_bands,
            "review_queue": self.review_queue[:50],
        }


def _cell(row: tuple, i: int) -> str:
    v = row[i] if i < len(row) else None
    return "" if v is None else str(v).strip()


def _to_float(text: str) -> float | None:
    try:
        return float(text)
    except (TypeError, ValueError):
        return None


def _material_from_name(name: str) -> str | None:
    import re
    m = re.search(r"(Nd|Yb|Er|Tm)\s*[:：]\s*([A-Za-z0-9]+)", name)
    if m:
        return f"{m.group(1)}:{m.group(2)}"
    for token in ("LBO", "KTP", "BBO", "BIBO", "KDP", "CaF2", "YAG", "YLF", "YVO4"):
        if token in name.upper():
            return token
    return None


def _build_item(
    db: Session,
    report: ImportReport,
    *,
    name: str,
    s1: str,
    s2: str,
    geometry_cells: tuple[str, ...],
    quantity: str,
    location: str,
    keeper: str,
    vendor: str,
    source_file: str,
) -> None:
    p1 = parse_coating(s1, "S1")
    p2 = parse_coating(s2, "S2")
    geo = parse_geometry(*geometry_cells)

    notes = [*p1.notes, *p2.notes, *geo.notes]
    confidences = {p1.confidence, p2.confidence}
    if "needs_review" in confidences and not (p1.bands or p2.bands):
        confidence = "needs_review"
    elif "needs_review" in confidences or "partial" in confidences:
        confidence = "partial"
    else:
        confidence = "parsed"

    condition = "ok"
    if any(marker in notes for marker in DAMAGE_MARKERS):
        condition = "damaged"
    elif any("未知" in n or "不能确定" in n or "未能确定" in n for n in notes):
        condition = "uncertain"

    raw_spec = " | ".join(filter(None, [f"S1:{s1}" if s1 else "", f"S2:{s2}" if s2 else "",
                                        *geometry_cells]))
    item = InventoryItem(
        category=classify_category(name),
        name=name,
        diameter_mm=geo.diameter_mm,
        roc_mm=geo.roc_mm,
        roc_is_flat=geo.roc_is_flat,
        thickness_mm=geo.thickness_mm,
        dimensions=geo.dimensions,
        cut_angle_theta_deg=geo.cut_angle_theta_deg,
        cut_angle_phi_deg=geo.cut_angle_phi_deg,
        cut_axis=geo.cut_axis,
        doping_pct=geo.doping_pct,
        material=_material_from_name(name),
        # a genuine 0 must survive; only a blank/unparseable cell defaults to 1
        quantity=1.0 if _to_float(quantity) is None else _to_float(quantity),
        location=location or None,
        keeper=keeper or None,
        vendor=vendor or None,
        raw_spec=raw_spec[:2000],
        parse_confidence=confidence,
        parse_notes=notes,
        condition=condition,
        source_file=source_file,
    )
    db.add(item)
    db.flush()
    for band in (*p1.bands, *p2.bands):
        db.add(CoatingSpec(
            item_id=item.id,
            surface=band.surface,
            wl_min_nm=band.wl_min_nm,
            wl_max_nm=band.wl_max_nm,
            function=band.function,
            value_type=band.value_type,
            comparator=band.comparator,
            value_pct=band.value_pct,
            raw_fragment=band.raw_fragment,
        ))
        report.coating_bands += 1

    report.imported += 1
    if confidence == "needs_review":
        report.needs_review += 1
        report.review_queue.append({"item_id": item.id, "name": name, "raw": raw_spec[:160], "notes": notes})
    elif confidence == "partial":
        report.partial += 1
        report.review_queue.append({"item_id": item.id, "name": name, "raw": raw_spec[:160], "notes": notes})


def import_workbook(db: Session, path: str | Path, *, source_file: str | None = None) -> ImportReport:
    """Import the lab inventory workbook.  Returns an :class:`ImportReport`.

    Requires ``openpyxl``; raises ``RuntimeError`` with a clear message if the
    dependency is missing.
    """
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("Inventory import requires the 'openpyxl' package") from exc

    path = Path(path)
    src = source_file or path.name
    wb = openpyxl.load_workbook(path, data_only=True)
    report = ImportReport()

    # Sheet 1: mirrors (header row 0; forward-filled name/size groups)
    ws = wb.worksheets[0]
    last_name = ""
    last_size = ""
    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if idx == 0:
            continue
        report.total_rows += 1
        raw_name = _cell(row, 1)
        s1, s2 = _cell(row, 2), _cell(row, 3)
        roc = _cell(row, 5)
        # Test the RAW cells: with forward-fill, `name` is non-empty from the
        # second row on, so an interior blank separator row would otherwise be
        # imported as a phantom duplicate of the previous item.
        if not any([raw_name, s1, s2, roc, _cell(row, 4), _cell(row, 6)]):
            report.skipped_empty += 1
            continue
        name = raw_name or last_name
        size = _cell(row, 4) or last_size
        last_name, last_size = name, size
        _build_item(
            db, report,
            name=name or "未命名镜片",
            s1=s1, s2=s2,
            geometry_cells=(size, roc),
            quantity=_cell(row, 6), location=_cell(row, 7),
            keeper=_cell(row, 8), vendor=_cell(row, 9),
            source_file=src,
        )

    # Sheet 2: crystals (no header row)
    if len(wb.worksheets) > 1:
        for row in wb.worksheets[1].iter_rows(values_only=True):
            name = _cell(row, 0)
            if not name:
                report.skipped_empty += 1
                continue
            report.total_rows += 1
            _build_item(
                db, report,
                name=name,
                s1=_cell(row, 1), s2=_cell(row, 2),
                geometry_cells=(_cell(row, 3),),
                quantity=_cell(row, 5), location=_cell(row, 6),
                keeper=_cell(row, 7), vendor=_cell(row, 8),
                source_file=src,
            )

    return report
